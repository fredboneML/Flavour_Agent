"""
Export cluster assignments for all algorithms used in erdbeere_v2_all_algorithms_mds.png
to an Excel file.

Output: outputs/cluster_assignments_all_algorithms.xlsx
  - One row per algorithm
  - Columns: Algorithm, Cluster_1 (newline-separated recipe names), Cluster_2, ...
"""

import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import numpy as np
import pandas as pd
from scipy.cluster.hierarchy import linkage, fcluster
from scipy.spatial.distance import squareform
from sklearn.manifold import MDS
from sklearn.preprocessing import normalize
from sklearn.metrics import silhouette_score, davies_bouldin_score, calinski_harabasz_score
from sklearn.cluster import KMeans, DBSCAN, SpectralClustering
from sklearn.mixture import GaussianMixture
from sklearn.decomposition import PCA
from pathlib import Path
from collections import Counter
import warnings
warnings.filterwarnings('ignore')

try:
    from sklearn.cluster import HDBSCAN as _HDBSCAN
    _has_hdbscan = True
except ImportError:
    _has_hdbscan = False

# ── Paths ─────────────────────────────────────────────────────────────────────
CSV_PATH    = Path('data/gold/Third_Trial_Set_PDM Erdbeere Gesamt 8-5-2026.csv')
IGNORE_PATH = Path('data/gold/ignone_substances.csv')
OUT_PATH    = Path('outputs/cluster_assignments_all_algorithms.xlsx')

OT1           = 'Odour-Type 1'
OT2           = 'Odour-Type 2'
OT3           = 'Odour-Type 3'
THRESHOLD_COL = 'Threshold'
REZ_COL       = 'Rez.-Nr.'
IDENT_COL     = 'Ident'
CAS_COL       = 'CAS-Nr.'
NAME_COL      = 'Name'
TOTAL_COL     = 'Totalmenge'

# ── Helpers ───────────────────────────────────────────────────────────────────
def to_float(v, fallback=0.0):
    if v is None: return fallback
    if isinstance(v, (int, float)): return float(v)
    try: return float(str(v).strip().replace(',', '.'))
    except: return fallback

def norm_term(term):
    if pd.isna(term) or not isinstance(term, str): return None
    t = term.lower().strip().replace('"', '').replace("'", '').rstrip('.,;:')
    return t if len(t) >= 2 else None

def build_vocabulary(df, feature_cols):
    all_terms = set()
    for col in feature_cols:
        if col in df.columns:
            for t in df[col].dropna().map(norm_term):
                if t: all_terms.add(t)
    return sorted(all_terms)

def build_recipe_vectors(df, recipes, feature_cols_weighted, use_threshold):
    feature_cols = [col for col, _ in feature_cols_weighted]
    vocab        = build_vocabulary(df, feature_cols)
    vocab_to_idx = {t: i for i, t in enumerate(vocab)}
    vectors      = np.zeros((len(recipes), len(vocab)), dtype=np.float64)
    for r_idx, recipe in enumerate(recipes):
        for _, row in df[df[REZ_COL] == recipe].iterrows():
            qty = float(row[TOTAL_COL])
            if qty <= 0: continue
            t_fac = 1.0  # no threshold
            ingr_base = qty * t_fac
            for col, col_weight in feature_cols_weighted:
                term = norm_term(row.get(col))
                if term and term in vocab_to_idx:
                    vectors[r_idx, vocab_to_idx[term]] += col_weight * ingr_base
    return vocab, normalize(vectors)

def cosine_dissimilarity(vecs):
    sim = np.clip(vecs @ vecs.T, -1.0, 1.0)
    diss = 1.0 - sim
    np.fill_diagonal(diss, 0.0)
    return diss

def resolve_noise(raw, dist_matrix):
    labels = raw.copy()
    noise  = np.where(labels == -1)[0]
    if len(noise) == 0: return labels + 1
    non_noise = np.where(labels != -1)[0]
    if len(non_noise) == 0: return np.ones(len(labels), dtype=int)
    for i in noise:
        labels[i] = labels[non_noise[np.argmin(dist_matrix[i, non_noise])]]
    return labels + 1

def kmedoids(dist_matrix, k, max_iter=500, random_state=42):
    n   = dist_matrix.shape[0]
    rng = np.random.RandomState(random_state)
    # Convert to Python int immediately so equality check is type-safe
    medoids = [int(x) for x in rng.choice(n, k, replace=False)]
    for _ in range(max_iter):
        labels = np.argmin(dist_matrix[:, medoids], axis=1)
        new_medoids = []
        for c in range(k):
            pts = np.where(labels == c)[0]
            if len(pts) == 0:
                new_medoids.append(medoids[c])
                continue
            sub = dist_matrix[np.ix_(pts, pts)]
            new_medoids.append(int(pts[np.argmin(sub.sum(axis=1))]))
        if new_medoids == medoids:
            break
        medoids = new_medoids
    return np.argmin(dist_matrix[:, medoids], axis=1) + 1

def fuzzy_cmeans(X, c, m=2.0, max_iter=500, tol=1e-7, random_state=42):
    rng = np.random.RandomState(random_state)
    n   = X.shape[0]
    U   = rng.dirichlet(np.ones(c), n).T
    for _ in range(max_iter):
        Um        = U ** m
        centroids = (Um @ X) / Um.sum(axis=1, keepdims=True)
        dist      = np.array([[np.linalg.norm(X[j] - centroids[i]) for j in range(n)] for i in range(c)])
        dist      = np.maximum(dist, 1e-12)
        ratio     = dist[None, :, :] / dist[:, None, :]
        U_new     = 1.0 / (ratio ** (2.0 / (m - 1))).sum(axis=1)
        U_new    /= U_new.sum(axis=0, keepdims=True)
        if np.max(np.abs(U_new - U)) < tol: break
        U = U_new
    return np.argmax(U, axis=0) + 1

def som_cluster(X, k, lr=0.5, sigma=1.5, n_iter=10000, random_state=42):
    rng    = np.random.RandomState(random_state)
    g      = int(np.ceil(np.sqrt(k)))
    n_nodes = g * g
    weights = normalize(rng.randn(n_nodes, X.shape[1]))
    node_pos = np.array([(r, c) for r in range(g) for c in range(g)], dtype=float)
    for t in range(n_iter):
        frac   = 1.0 - t / n_iter
        lr_t   = lr * frac
        sig_t  = max(sigma * frac, 0.01)
        xi     = X[rng.randint(0, X.shape[0])]
        bmu    = np.argmin(np.linalg.norm(weights - xi, axis=1))
        d2     = ((node_pos - node_pos[bmu]) ** 2).sum(axis=1)
        h      = np.exp(-d2 / (2 * sig_t ** 2))
        weights += lr_t * h[:, None] * (xi - weights)
    node_labels = KMeans(n_clusters=k, n_init=30, random_state=random_state).fit_predict(weights)
    recipe_labels = np.array([node_labels[np.argmin(np.linalg.norm(weights - xi, axis=1))]
                               for xi in X])
    return recipe_labels + 1

def dec_simplified(X, n_clusters, enc_dim=8, n_iter=300, random_state=42):
    d   = min(enc_dim, X.shape[1], X.shape[0] - 1)
    Z   = PCA(n_components=d, random_state=random_state).fit_transform(X)
    centers = KMeans(n_clusters=n_clusters, n_init=50, random_state=random_state).fit(Z).cluster_centers_.copy()
    for _ in range(n_iter):
        dist2 = np.sum((Z[:, None, :] - centers[None, :, :]) ** 2, axis=2)
        q     = 1.0 / (1.0 + dist2)
        q    /= q.sum(axis=1, keepdims=True)
        f     = q.sum(axis=0)
        p     = (q ** 2) / np.maximum(f, 1e-12)
        p    /= p.sum(axis=1, keepdims=True)
        for j in range(n_clusters):
            centers[j] = (p[:, j:j+1] * Z).sum(axis=0) / max(p[:, j].sum(), 1e-12)
    return np.argmax(q, axis=1) + 1

# ── Load data ─────────────────────────────────────────────────────────────────
print('Loading data...')
df_raw = pd.read_csv(CSV_PATH, dtype=str)
df_raw[TOTAL_COL]     = df_raw[TOTAL_COL].apply(to_float)
df_raw[THRESHOLD_COL] = df_raw[THRESHOLD_COL].apply(to_float)
df = df_raw[df_raw[REZ_COL].notna()].copy()

if IGNORE_PATH.exists():
    ign        = pd.read_csv(IGNORE_PATH)
    ign_idents = set(ign[IDENT_COL].dropna().astype(str).str.strip())
    names_ign  = {str(n).lower().strip() for n in ign[NAME_COL]}
    mask = (
        df[IDENT_COL].astype(str).str.strip().isin(ign_idents) |
        df[NAME_COL].str.lower().str.strip().isin(names_ign)
    )
    cas_ign = set(df.loc[mask, CAS_COL].dropna().astype(str).str.strip())
    df.loc[df[CAS_COL].astype(str).str.strip().isin(cas_ign), TOTAL_COL] = 0.0

per_recipe_total = df.groupby(REZ_COL)[TOTAL_COL].transform('sum')
df[TOTAL_COL] = np.where(per_recipe_total > 0,
                          df[TOTAL_COL] / per_recipe_total, df[TOTAL_COL])

recipes = df[REZ_COL].unique().tolist()
n       = len(recipes)
print(f'Recipes: {n}')

# ── Build M1 vectors & dissimilarity ─────────────────────────────────────────
print('Building feature vectors...')
_, vecs_m2 = build_recipe_vectors(df, recipes, [(OT1, 1.0)], use_threshold=False)
diss_m2    = cosine_dissimilarity(vecs_m2)

# ── Determine OPTIMAL_K (majority vote, identical to notebook) ────────────────
print('Determining optimal k...')
k_range = range(2, min(n - 1, 9))
Z_m2_sq = linkage(squareform(diss_m2, checks=False), method='ward')

merge_heights = Z_m2_sq[:, 2]
height_gaps   = np.diff(merge_heights[::-1])
best_k_ward   = int(np.argmax(height_gaps) + 2)

sil_scores, db_scores, ch_scores = [], [], []
for k in k_range:
    lbl = fcluster(Z_m2_sq, t=k, criterion='maxclust')
    if len(np.unique(lbl)) < 2:
        sil_scores.append(-1); db_scores.append(9999); ch_scores.append(0); continue
    sil_scores.append(silhouette_score(diss_m2, lbl, metric='precomputed'))
    db_scores.append(davies_bouldin_score(vecs_m2, lbl))
    ch_scores.append(calinski_harabasz_score(vecs_m2, lbl))

best_k_sil = int(list(k_range)[int(np.argmax(sil_scores))])
best_k_db  = int(list(k_range)[int(np.argmin(db_scores))])
best_k_ch  = int(list(k_range)[int(np.argmax(ch_scores))])

votes = [best_k_ward, best_k_sil, best_k_db, best_k_ch]
vote_counts = Counter(votes)
OPTIMAL_K   = vote_counts.most_common(1)[0][0]
K           = OPTIMAL_K
print(f'OPTIMAL_K = {K}')

# ── Ward labels ───────────────────────────────────────────────────────────────
labels_ward = fcluster(Z_m2_sq, t=K, criterion='maxclust')

# ── DBSCAN ────────────────────────────────────────────────────────────────────
print('Running DBSCAN...')
best_db_eps, best_db_raw, best_db_diff = None, None, 999
for eps in np.arange(0.01, 1.00, 0.01):
    raw     = DBSCAN(eps=eps, min_samples=2, metric='precomputed').fit_predict(diss_m2)
    n_clust = len(set(raw) - {-1})
    diff    = abs(n_clust - K)
    # Prefer solutions with more clusters (closer to K) over degenerate 1-cluster
    if diff < best_db_diff or (diff == best_db_diff and n_clust > len(set(best_db_raw) - {-1})):
        best_db_eps, best_db_raw, best_db_diff = eps, raw, diff
    if n_clust == K and (raw == -1).sum() == 0:
        best_db_eps, best_db_raw, best_db_diff = eps, raw, diff
        break
labels_dbscan   = resolve_noise(best_db_raw, diss_m2)
dbscan_n_clust  = len(set(best_db_raw) - {-1})
dbscan_n_noise  = (best_db_raw == -1).sum()
print(f'  DBSCAN best eps={best_db_eps:.2f} → {dbscan_n_clust} raw clusters, {dbscan_n_noise} noise pts resolved')

# ── HDBSCAN ───────────────────────────────────────────────────────────────────
print('Running HDBSCAN...')
if _has_hdbscan:
    best_hdb_mcs, best_hdb_raw = None, None
    for mcs in range(2, n + 1):
        raw     = _HDBSCAN(min_cluster_size=mcs, metric='precomputed').fit_predict(diss_m2)
        n_clust = len(set(raw) - {-1})
        if best_hdb_mcs is None or abs(n_clust - K) < abs(len(set(best_hdb_raw) - {-1}) - K):
            best_hdb_mcs, best_hdb_raw = mcs, raw
    labels_hdbscan = resolve_noise(best_hdb_raw, diss_m2)
else:
    labels_hdbscan = None

# ── k-Means ───────────────────────────────────────────────────────────────────
print('Running k-Means...')
labels_kmeans = KMeans(n_clusters=K, n_init=100, random_state=42).fit_predict(vecs_m2) + 1

# ── k-Medoids ─────────────────────────────────────────────────────────────────
print('Running k-Medoids...')
labels_kmedoids = kmedoids(diss_m2, k=K)

# ── GMM ───────────────────────────────────────────────────────────────────────
print('Running GMM...')
enc_dim_gmm  = min(K, vecs_m2.shape[0] - 1, vecs_m2.shape[1])
Z_gmm        = PCA(n_components=enc_dim_gmm, random_state=42).fit_transform(vecs_m2)
labels_gmm   = GaussianMixture(n_components=K, n_init=30, covariance_type='full',
                                random_state=42).fit_predict(Z_gmm) + 1

# ── Spectral ──────────────────────────────────────────────────────────────────
print('Running Spectral...')
affinity         = np.clip(1.0 - diss_m2, 0.0, 1.0); np.fill_diagonal(affinity, 1.0)
labels_spectral  = SpectralClustering(n_clusters=K, affinity='precomputed',
                                       n_init=100, random_state=42).fit_predict(affinity) + 1

# ── Fuzzy c-Means ─────────────────────────────────────────────────────────────
print('Running Fuzzy c-Means...')
enc_dim_fcm  = min(max(K, 4), vecs_m2.shape[0] - 1, vecs_m2.shape[1])
Z_fcm        = PCA(n_components=enc_dim_fcm, random_state=42).fit_transform(vecs_m2)
labels_fcm   = fuzzy_cmeans(Z_fcm, c=K)

# ── SOM ───────────────────────────────────────────────────────────────────────
print('Running SOM...')
enc_dim_som  = min(8, vecs_m2.shape[0] - 1, vecs_m2.shape[1])
Z_som        = PCA(n_components=enc_dim_som, random_state=42).fit_transform(vecs_m2)
labels_som   = som_cluster(Z_som, k=K)

# ── DEC simplified ────────────────────────────────────────────────────────────
print('Running DEC...')
labels_dec   = dec_simplified(vecs_m2, n_clusters=K)

# ── Collect ───────────────────────────────────────────────────────────────────
algo_labels_dict = {
    'k-Means':       labels_kmeans,
    'k-Medoids':     labels_kmedoids,
    'Ward (M2)':     labels_ward,
    'DBSCAN':        labels_dbscan,
    'HDBSCAN':       labels_hdbscan,
    'GMM':           labels_gmm,
    'Spectral':      labels_spectral,
    'Fuzzy c-Means': labels_fcm,
    'SOM':           labels_som,
    'DEC (simpl.)':  labels_dec,
}

# ── Build Excel ───────────────────────────────────────────────────────────────
print('Building Excel...')
recipes_arr = np.array(recipes)

# Notes for algorithms with known caveats
algo_notes = {
    'DBSCAN': (f'No eps yields exactly k={K} without noise. '
               f'Best eps={best_db_eps:.2f} → {dbscan_n_clust} raw clusters, '
               f'{dbscan_n_noise} noise pts reassigned to nearest cluster.'),
}

rows = []
for algo_name, lbl in algo_labels_dict.items():
    if lbl is None:
        rows.append({'Algorithm': algo_name, 'Cluster_1': 'N/A',
                     'Cluster_2': 'N/A', 'Notes': 'Not available (sklearn < 1.3)'})
        continue
    unique_clusters = sorted(np.unique(lbl))
    row = {'Algorithm': algo_name}
    for cid in unique_clusters:
        mask = lbl == cid
        recipe_list = sorted(recipes_arr[mask].tolist())
        row[f'Cluster_{cid}'] = '\n'.join(recipe_list)
    counts = {cid: int((lbl == cid).sum()) for cid in unique_clusters}
    count_str = '  |  '.join(f'C{c}: {n}' for c, n in counts.items())
    row['Notes'] = algo_notes.get(algo_name, '') or f'k={len(unique_clusters)}  ({count_str})'
    rows.append(row)

df_out = pd.DataFrame(rows)

# Ensure all cluster columns exist and are in order
cluster_cols = sorted([c for c in df_out.columns if c.startswith('Cluster_')],
                      key=lambda x: int(x.split('_')[1]))
df_out = df_out[['Algorithm'] + cluster_cols + ['Notes']].fillna('')

with pd.ExcelWriter(OUT_PATH, engine='openpyxl') as writer:
    df_out.to_excel(writer, index=False, sheet_name='Cluster Assignments')
    ws = writer.sheets['Cluster Assignments']

    # Style: wrap text, wider columns, bold header
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill('solid', fgColor='2B6CB0')
    header_font = Font(bold=True, color='FFFFFF', size=11)

    for col_idx, col_name in enumerate(df_out.columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font    = header_font
        cell.fill    = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # Column widths & row heights
    ws.column_dimensions['A'].width = 18
    for col_idx in range(2, len(df_out.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 36

    for row_idx in range(2, len(df_out) + 2):
        ws.row_dimensions[row_idx].height = None  # auto
        for col_idx in range(1, len(df_out.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = Alignment(vertical='top', wrap_text=True)

    # Alternate row shading
    alt_fill = PatternFill('solid', fgColor='EBF5FB')
    for row_idx in range(2, len(df_out) + 2, 2):
        for col_idx in range(1, len(df_out.columns) + 1):
            ws.cell(row=row_idx, column=col_idx).fill = alt_fill

print(f'Done. Saved to: {OUT_PATH}')
print()
print('Summary:')
print(df_out[['Algorithm'] + [c for c in cluster_cols]].to_string(index=False))
