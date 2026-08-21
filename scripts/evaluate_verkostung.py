"""First-glance evaluation of all scoring methods against the 25-06-2026 Verkostung panel.

Recomputes the 15 scoring variants (identical to notebooks/scoring_ms_assessment.ipynb) on the
new all-fruit PDM export, then scores each method against the panel ground truth in
`Ergebnisse Verkostung 25_06_2026.xlsx`, using two truth definitions:
  Method 1 = M1 Label Propagation (column B),  Method 2 = M2 Rule Based (column C).

True-label rule per recipe (Prozent % = column H):
  - H >= 60 : true label = the single value in B (method 1) / C (method 2).
  - H <  60 : acceptable labels = tokens in D (Cluster vorgegeben) and E (Jan Free Sorting)
              that are literally one of the 7 cluster names; if none, the recipe is ignored.
A prediction is correct if it equals ANY label in the recipe's acceptable set (lenient match).

Output: outputs/verkostung_eval_25_06_2026.xlsx

Usage:
    python3 scripts/evaluate_verkostung.py
"""

import re
from pathlib import Path

import openpyxl
import pandas as pd

from scripts.ms_scoring import (
    load_weight_matrix,
    load_recipes,
    load_rohstoffe_weights,
    load_avg_meli,
    compute_scores,
    assign_clusters,
    map_to_panel,
    apply_zscore_quantities,
    central_amount_when_present,
    apply_per_fruit_zscore,
    apply_squared_zscore_quantities,
    rescale_warm_weights,
    compute_tfidf_scores,
    compute_topk_scores,
    compute_cosine_scores,
    ensemble_scores,
)
from scripts.odor_threshold import (
    cas_odor_sets, masked_weights, threshold_variants, cas_threshold, factor_frame,
)

ROOT = Path(__file__).resolve().parent.parent
GOLD = ROOT / "data" / "gold"

SCORING_XLSX = GOLD / "Scoring Index_Beispielrechung.xlsx"
PDM_CSV = GOLD / "PDM_Rezepturen_Gesamt_nurP_18_08_2026.csv"
PDM_XLSX = GOLD / "PDM_Rezepturen_Gesamt_nurP_18_08_2026.xlsx"


def fruit_labels() -> dict:
    """Rez.-Nr. -> fruit type = leading token of Rezepturbezeichnung (from the source XLSX)."""
    df = pd.read_excel(PDM_XLSX, sheet_name="Rezept", header=14, dtype=str)
    fruit = df["Rezepturbezeichnung"].astype(str).str.split(r"[\s\-]", n=1).str[0].str.strip()
    return dict(zip(df["Rez.-Nr."].astype(str).str.strip(), fruit))


# ── Human-readable explanation of the key methods (hand-over reference sheet) ──────────
# Common notation: amount = per-recipe-normalised ingredient amount (a recipe's amounts sum to 1);
# z = amount / D where D is the "typical amount" denominator; score[c] = Σ over ingredients of
# z × weight[CAS][c]; the recipe is assigned to argmax over clusters c. weight = Melanie's expert
# 0–20 matrix. Denominators D: AvgMeli = mean amount when present over Melanie's 9 curated reference
# recipes; all-fruit mean/median = over all 3,981 recipes; per-fruit mean/median = over recipes of
# the same fruit (Rezepturbezeichnung), with an all-fruit fallback for sparse fruits.
KEY_METHODS_EXPLANATION = [
    {"Method": "MS Z-Score",
     "Formula": "z = amount / AvgMeli;  score[c] = Σ z × weight[CAS][c];  predict = argmax",
     "Explanation & Idea behind it": "Champion baseline. Divide each ingredient's amount by its typical dosage (AvgMeli = mean over Melanie's 9 curated strawberry reference recipes), multiply by the expert cluster weights, pick the top-scoring cluster. Idea: a recipe belongs to the cluster whose signature ingredients it over-uses relative to normal.",
     "Example": "A recipe with 3× the usual Furaneol (a 'warm' marker) gets a high warm z-score → predicted warm."},

    {"Method": "Z-Score mean (all fruits)",
     "Formula": "z = amount / mean_all_fruits(amount when present)",
     "Explanation & Idea behind it": "Same as MS Z-Score but the typical amount is the MEAN over all 3,981 recipes (all fruit types) instead of Melanie's 9 reference recipes. Idea: derive 'typical' from the whole dataset. Worse in practice — the all-fruit average is noisier.",
     "Example": "Furaneol's typical amount is averaged across strawberry, apple, cherry… diluting its strawberry signature."},
    {"Method": "Z-Score (per-fruit mean)",
     "Formula": "z = amount / mean_same_fruit(amount when present)",
     "Explanation & Idea behind it": "Same as Z-Score mean but the mean is computed WITHIN each recipe's own fruit type (all-fruit fallback for sparse fruits). Idea: judge 'unusually much' against that fruit's own norms, not a blended average.",
     "Example": "For a strawberry recipe, Furaneol's typical amount is the mean over strawberry recipes only."},

    {"Method": "Z-Score median (all fruits)",
     "Formula": "z = amount / median_all_fruits(amount when present)",
     "Explanation & Idea behind it": "Same as Z-Score mean but uses the MEDIAN (robust to outlier recipes) over all fruits.",
     "Example": "A few recipes dosing Furaneol very high no longer inflate its 'typical' amount."},
    {"Method": "Z-Score (per-fruit median)",
     "Formula": "z = amount / median_same_fruit(amount when present)",
     "Explanation & Idea behind it": "Same but over each recipe's own fruit (robust + fruit-specific). Best of the data-derived denominators on the strawberry test.",
     "Example": "Strawberry-only median of Furaneol as the denominator for a strawberry recipe."},

    {"Method": "Odor-gate",
     "Formula": "w'[CAS][c] = weight[CAS][c] if c matches the ingredient's odor type else 0;  z = amount / AvgMeli;  score[c] = Σ z × w'[CAS][c]",
     "Explanation & Idea behind it": "MS Z-Score, but before scoring, zero any ingredient→cluster weight whose cluster disagrees with the ingredient's odor type (Odour-Type 1/2/3, union). Idea: an ingredient should only vote for clusters it actually smells like. Best method overall (70.6 / 55.0).",
     "Example": "2-Methylbuttersäure (odor fruity/unpleasant/dairy) keeps its unpleasant & citrus weights, loses warm & exotic."},
    {"Method": "Odor-gate (per-fruit mean)",
     "Formula": "same odor gate, z = amount / mean_same_fruit",
     "Explanation & Idea behind it": "The odor-gate but the z-score denominator is the per-fruit MEAN instead of Melanie's AvgMeli. Its non-per-fruit variant is 'Odor-gate' above (AvgMeli).",
     "Example": "Strawberry recipe scored with strawberry-mean z, then odor-gated weights."},
    {"Method": "Odor-gate (per-fruit median)",
     "Formula": "same odor gate, z = amount / median_same_fruit",
     "Explanation & Idea behind it": "Same as Odor-gate (per-fruit mean) but with the per-fruit MEDIAN denominator (robust).",
     "Example": "Strawberry-only median z with odor-gated weights."},

    {"Method": "inv-threshold a=0.25",
     "Formula": "z = amount / AvgMeli;  f = (median_threshold / threshold)^0.25;  score[c] = Σ z × f × weight[CAS][c]",
     "Explanation & Idea behind it": "MS Z-Score with each ingredient additionally weighted by potency — the fourth root of inverse odor threshold (low threshold ppm = potent). α=0.25 is a gentle nudge; it ties the champion, stronger α hurts (potency is already in the weights).",
     "Example": "A very potent trace ester (tiny threshold) gets a mild boost toward its cluster."},
    {"Method": "inv-threshold a=0.25 (per-fruit mean)",
     "Formula": "z = amount / mean_same_fruit;  f = (median_threshold / threshold)^0.25",
     "Explanation & Idea behind it": "Same potency weighting but the z-score denominator is the per-fruit MEAN. Non-per-fruit variant: 'inv-threshold a=0.25' (AvgMeli).",
     "Example": "Per-fruit-mean z, then multiplied by the fourth-root potency factor."},
    {"Method": "inv-threshold a=0.25 (per-fruit median)",
     "Formula": "z = amount / median_same_fruit;  f = (median_threshold / threshold)^0.25",
     "Explanation & Idea behind it": "Same but per-fruit MEDIAN denominator. Among the strongest per-fruit variants (ties champion on M1, edges it on M2).",
     "Example": "Per-fruit-median z × potency factor."},

    {"Method": "Odor-gate + inv-threshold a=0.25",
     "Formula": "odor-gated weights w' + potency f=(median_threshold/threshold)^0.25;  z = amount / AvgMeli",
     "Explanation & Idea behind it": "Combine the odor-gate (weights pruned by odor type) with the α=0.25 potency weighting.",
     "Example": "Odor-gated weights, ingredient contributions scaled by the fourth-root potency factor."},
    {"Method": "Odor-gate + inv-threshold a=0.25 (per-fruit mean)",
     "Formula": "same, z = amount / mean_same_fruit",
     "Explanation & Idea behind it": "Same as above but per-fruit MEAN z-denominator. Non-per-fruit variant: 'Odor-gate + inv-threshold a=0.25'.",
     "Example": "Per-fruit-mean z, odor-gated weights, α=0.25 potency."},
    {"Method": "Odor-gate + inv-threshold a=0.25 (per-fruit median)",
     "Formula": "same, z = amount / median_same_fruit",
     "Explanation & Idea behind it": "Same but per-fruit MEDIAN. Ties the champion on M1 and edges it on M2 (64.7 / 55.0).",
     "Example": "Per-fruit-median z, odor-gated weights, α=0.25 potency."},

    {"Method": "Odor-gate + inv-threshold a=0.5",
     "Formula": "odor-gated weights w' + potency f=(median_threshold/threshold)^0.5;  z = amount / AvgMeli",
     "Explanation & Idea behind it": "Same as the α=0.25 combination but stronger potency weighting (square root instead of fourth root).",
     "Example": "Odor-gated weights, contributions scaled by the square-root potency factor."},
    {"Method": "Odor-gate + inv-threshold a=0.5 (per-fruit mean)",
     "Formula": "same, z = amount / mean_same_fruit",
     "Explanation & Idea behind it": "Same but per-fruit MEAN z-denominator. Non-per-fruit variant: 'Odor-gate + inv-threshold a=0.5'.",
     "Example": "Per-fruit-mean z, odor-gated weights, α=0.5 potency."},
    {"Method": "Odor-gate + inv-threshold a=0.5 (per-fruit median)",
     "Formula": "same, z = amount / median_same_fruit",
     "Explanation & Idea behind it": "Same but per-fruit MEDIAN z-denominator.",
     "Example": "Per-fruit-median z, odor-gated weights, α=0.5 potency."},
]
IGNORE_PATH = GOLD / "ignone_substances.csv"
VERK_XLSX = GOLD / "Ergebnisse Verkostung 25_06_2026.xlsx"
OUT_XLSX = ROOT / "outputs" / "verkostung_eval_25_06_2026.xlsx"

# MS cluster name -> panel/expert cluster name (copied from notebook cell 2).
MS_TO_PANEL = {
    "warm": "warm",
    "Unpleasant": "unpleasant",
    "green": "green",
    "floral": "floral",
    "citrus": "fruity",
    # "exotic" and "Outlayer" have no panel equivalent -> map() yields NaN
}

# The 7 expert/panel cluster names (lowercase) used as the truth vocabulary.
SEVEN = {"warm", "floral", "walderdbeere", "green", "dairy", "unpleasant", "fruity"}
# Clusters a z-method can actually output (via MS_TO_PANEL).
REACHABLE = {"warm", "unpleasant", "green", "floral", "fruity"}

Z_METHODS = {
    "MS Z-Score", "Z-Score mean", "Z-Score median", "Warm + Z-Score",
    "Squared Z-Score", "Warm + Sq.Z-Score", "Ensemble orig+Z", "Ensemble warm+Z",
}


def base_id(x: object) -> str:
    """Strip a trailing 'P' variant marker for cross-file recipe matching."""
    return re.sub(r"P$", "", str(x).strip())


def norm_label(v: object) -> str:
    """Lowercase, strip whitespace, drop '(...)' suffixes, map German 'grün' -> 'green'."""
    if v is None:
        return ""
    s = str(v).split("(")[0].strip().lower()
    return "green" if s == "grün" else s


def split_tokens(v: object) -> list[str]:
    """Split a D/E cell on '/' and '-' into normalized tokens."""
    if v is None:
        return []
    return [norm_label(t) for t in re.split(r"[/\-]", str(v)) if t.strip()]


# ── 1. Recompute the 15 method score matrices (mirrors notebook cell 37) ──────
def build_variants():
    weights = load_weight_matrix(SCORING_XLSX)
    recipes_df = load_recipes(PDM_CSV, IGNORE_PATH)
    avg_meli = load_avg_meli(SCORING_XLSX, csv_path=PDM_CSV)
    weights_rohst = load_rohstoffe_weights(SCORING_XLSX)
    weights_warm = rescale_warm_weights(weights)

    scores = compute_scores(recipes_df, weights)
    scores_z = compute_scores(apply_zscore_quantities(recipes_df, avg_meli), weights)
    # Data-derived Z-Scores: divide amounts by the mean / median amount-when-present per CAS,
    # computed from the recipe dataset itself (median is robust to outlier recipes).
    mean_present = central_amount_when_present(recipes_df, "mean")
    median_present = central_amount_when_present(recipes_df, "median")
    scores_z_mean = compute_scores(apply_zscore_quantities(recipes_df, mean_present), weights)
    scores_z_median = compute_scores(apply_zscore_quantities(recipes_df, median_present), weights)
    scores_rohst = compute_scores(recipes_df, weights_rohst)
    sc_warm = compute_scores(recipes_df, weights_warm)
    sc_warm_z = compute_scores(apply_zscore_quantities(recipes_df, avg_meli), weights_warm)
    recipes_z2 = apply_squared_zscore_quantities(recipes_df, avg_meli)
    sc_z2 = compute_scores(recipes_z2, weights)
    sc_warm_z2 = compute_scores(recipes_z2, weights_warm)
    sc_tfidf = compute_tfidf_scores(recipes_df, weights)
    sc_tfidf_warm = compute_tfidf_scores(recipes_df, weights_warm)
    sc_top5 = compute_topk_scores(recipes_df, weights, k=5)
    sc_top5_warm = compute_topk_scores(recipes_df, weights_warm, k=5)
    sc_cosine = compute_cosine_scores(recipes_df, weights)
    sc_cosine_warm = compute_cosine_scores(recipes_df, weights_warm)
    sc_ens_oz = ensemble_scores(scores, scores_z)
    sc_ens_wz = ensemble_scores(sc_warm, sc_warm_z)

    variants = {
        "MS original": scores,
        "MS Z-Score": scores_z,
        "Z-Score mean": scores_z_mean,
        "Z-Score median": scores_z_median,
        "MS Rohstoffe": scores_rohst,
        "Warm rescale": sc_warm,
        "Warm + Z-Score": sc_warm_z,
        "Squared Z-Score": sc_z2,
        "Warm + Sq.Z-Score": sc_warm_z2,
        "TF-IDF": sc_tfidf,
        "TF-IDF + warm": sc_tfidf_warm,
        "Top-5": sc_top5,
        "Top-5 + warm": sc_top5_warm,
        "Cosine": sc_cosine,
        "Cosine + warm": sc_cosine_warm,
        "Ensemble orig+Z": sc_ens_oz,
        "Ensemble warm+Z": sc_ens_wz,
    }

    # ── Odor-type gate + odor-threshold (OAV) methods ────────────────────────────
    # Odor data covers the strawberry palette; keep_unknown=True leaves non-strawberry
    # ingredients ungated so all 3,981 recipes stay scoreable for later evaluation.
    z_base = apply_zscore_quantities(recipes_df, avg_meli)
    odor = cas_odor_sets()
    w_gate = masked_weights(weights, odor, keep_unknown=True)

    # Threshold combinations on plain MS Z-Score (expert weights)
    variants.update(threshold_variants(recipes_df, weights, z_base, weights, prefix=""))
    # Odor-type gate, alone and combined with each threshold transform
    variants["Odor-gate"] = compute_scores(z_base, w_gate)
    variants.update(threshold_variants(recipes_df, weights, z_base, w_gate, prefix="Odor-gate + "))

    # ── Per-fruit denominator variants of the top methods ────────────────────────
    # z-score divides amounts by the typical amount computed within each recipe's own fruit
    # (Rezepturbezeichnung), instead of over all fruits.
    fruit_of = fruit_labels()
    thr, med = cas_threshold()
    for stat in ("mean", "median"):
        zpf = apply_per_fruit_zscore(recipes_df, fruit_of, stat)
        tag = f" (per-fruit {stat})"
        variants[f"Z-Score{tag}"] = compute_scores(zpf, weights)
        variants[f"Odor-gate{tag}"] = compute_scores(zpf, w_gate)
        variants[f"inv-threshold a=0.25{tag}"] = compute_scores(
            factor_frame(zpf, thr, med, "inv", 0.25), weights)
        variants[f"Odor-gate + inv-threshold a=0.25{tag}"] = compute_scores(
            factor_frame(zpf, thr, med, "inv", 0.25), w_gate)
        variants[f"Odor-gate + inv-threshold a=0.5{tag}"] = compute_scores(
            factor_frame(zpf, thr, med, "inv", 0.5), w_gate)

    return variants


# ── 2. Build the truth table from the Verkostung file ─────────────────────────
def build_truth() -> pd.DataFrame:
    wb = openpyxl.load_workbook(VERK_XLSX, data_only=True)
    ws = wb["Tabelle1"]
    rows = []
    for r in range(3, ws.max_row + 1):
        rec = ws.cell(r, 1).value
        if rec is None:
            continue
        B, C, D, E, H = (ws.cell(r, c).value for c in (2, 3, 4, 5, 8))
        if H is None or not isinstance(H, (int, float)):
            continue
        if H >= 60:
            m1 = {norm_label(B)} - {""}
            m2 = {norm_label(C)} - {""}
        else:
            de = {t for t in (split_tokens(D) + split_tokens(E)) if t in SEVEN}
            m1 = m2 = de
        rows.append(
            {
                "Recipe": str(rec).strip(),
                "base": base_id(rec),
                "H": H,
                "B_M1": B,
                "C_M2": C,
                "D": D,
                "E": E,
                "M1_set": sorted(m1),
                "M2_set": sorted(m2),
            }
        )
    return pd.DataFrame(rows)


def accuracy(preds_by_method: dict, truth: pd.DataFrame, set_col: str) -> pd.DataFrame:
    """Per-method accuracy: prediction correct if in the recipe's acceptable label set."""
    evalset = truth[truth[set_col].map(len) > 0]
    out = []
    for method, panel_pred in preds_by_method.items():
        corr = tot = r_corr = r_tot = 0
        for _, row in evalset.iterrows():
            pid = row["pdm_id"]
            if pid is None:
                continue
            pred = panel_pred.get(pid)
            pred = None if (pred is None or (isinstance(pred, float))) else str(pred).lower()
            labels = set(row[set_col])
            hit = pred in labels
            tot += 1
            corr += hit
            if labels & REACHABLE:  # reachable-only view
                r_tot += 1
                r_corr += hit
        out.append(
            {
                "Method": method,
                "z_method": method in Z_METHODS,
                "Correct": corr,
                "Total": tot,
                "Accuracy_%": round(100 * corr / tot, 1) if tot else None,
                "Correct_reachable": r_corr,
                "Total_reachable": r_tot,
                "Accuracy_reachable_%": round(100 * r_corr / r_tot, 1) if r_tot else None,
            }
        )
    return pd.DataFrame(out).sort_values("Accuracy_%", ascending=False, ignore_index=True)


def per_cluster_accuracy(
    preds_by_method: dict, truth: pd.DataFrame, set_col: str, reachable_only: bool = False
) -> pd.DataFrame:
    """Accuracy per TRUE cluster × method: read down a method column to see which clusters it
    nails vs misses. Rows = true label (joined ``set_col``), columns = n + one % per method.

    reachable_only: drop recipes whose true cluster has no weight column (i.e. dairy /
    Walderdbeere — clusters the methods can never predict), so scores reflect only reachable cases.
    """
    evalset = truth[truth[set_col].map(len) > 0].copy()
    if reachable_only:
        evalset = evalset[evalset[set_col].map(lambda s: bool(set(s) & REACHABLE))].copy()
    evalset["_group"] = evalset[set_col].map(lambda s: ", ".join(s))
    totals = {m: 0 for m in preds_by_method}
    rows = []
    for group, g in evalset.groupby("_group"):
        row = {"True_cluster": group, "n": len(g)}
        for method, panel_pred in preds_by_method.items():
            corr = 0
            for _, r in g.iterrows():
                pred = panel_pred.get(r["pdm_id"])
                pred = None if (pred is None or isinstance(pred, float)) else str(pred).lower()
                corr += pred in set(r[set_col])
            totals[method] += corr
            row[method] = round(100 * corr / len(g), 0)
        rows.append(row)
    df = pd.DataFrame(rows).sort_values("n", ascending=False, ignore_index=True)
    # Weighted overall row: total recipes and each method's accuracy over all of them.
    n_total = len(evalset)
    total_row = {"True_cluster": "TOTAL", "n": n_total,
                 **{m: round(100 * totals[m] / n_total, 0) if n_total else None for m in preds_by_method}}
    return pd.concat([df, pd.DataFrame([total_row])], ignore_index=True)


def main() -> None:
    variants = build_variants()
    n_recipes = len(next(iter(variants.values())))
    print(f"Recomputed {len(variants)} methods over {n_recipes} recipes")

    # Raw MS argmax cluster + panel-mapped prediction per method.
    ms_cluster = {m: assign_clusters(sc) for m, sc in variants.items()}
    panel_pred = {m: map_to_panel(a, MS_TO_PANEL) for m, a in ms_cluster.items()}

    truth = build_truth()
    pdm_base = {base_id(x): str(x).strip() for x in ms_cluster["MS original"].index}
    truth["in_pdm"] = truth["base"].isin(pdm_base)
    truth["pdm_id"] = truth["base"].map(pdm_base)
    truth["evaluable"] = truth["in_pdm"] & (truth["M1_set"].map(len) > 0)

    n_eval = int(truth["evaluable"].sum())
    print(f"Verkostung recipes: {len(truth)}  |  in PDM: {int(truth['in_pdm'].sum())}  |  evaluable subset = {n_eval}")

    eval_truth = truth[truth["evaluable"]].copy()
    acc_m1 = accuracy(panel_pred, eval_truth, "M1_set")
    acc_m2 = accuracy(panel_pred, eval_truth, "M2_set")
    # Method order, best-first (M1 reachable, then M2 reachable, then M1 raw) — reused across sheets.
    _rank = acc_m1[["Method", "Accuracy_reachable_%", "Accuracy_%"]].merge(
        acc_m2[["Method", "Accuracy_reachable_%"]].rename(
            columns={"Accuracy_reachable_%": "_m2r"}), on="Method")
    method_order = _rank.sort_values(
        ["Accuracy_reachable_%", "_m2r", "Accuracy_%"], ascending=False)["Method"].tolist()

    # Recipes whose truth set has no reachable cluster (only dairy/Walderdbeere) — the
    # z-methods can never predict these, so they are excluded from the reachable-only view.
    def _unreachable(labels: list[str]) -> bool:
        return len(labels) > 0 and not (set(labels) & REACHABLE)

    excluded_rows = []
    for _, row in eval_truth.iterrows():
        for m, col in (("M1", "M1_set"), ("M2", "M2_set")):
            if _unreachable(row[col]):
                excluded_rows.append(
                    {"Recipe": row["Recipe"], "truth_method": m,
                     "H": row["H"], "truth": ", ".join(row[col])}
                )
    excluded_df = pd.DataFrame(excluded_rows)
    n_keep_m1 = n_eval - int((excluded_df["truth_method"] == "M1").sum()) if len(excluded_df) else n_eval
    n_keep_m2 = n_eval - int((excluded_df["truth_method"] == "M2").sum()) if len(excluded_df) else n_eval

    # Subset predictions: one row per evaluable recipe, MS cluster + panel pred per method.
    subset_rows = []
    for _, row in eval_truth.iterrows():
        pid = row["pdm_id"]
        rec = {"Recipe": row["Recipe"], "H": row["H"],
               "M1_set": ", ".join(row["M1_set"]), "M2_set": ", ".join(row["M2_set"])}
        for m in variants:
            rec[f"{m} [MS]"] = ms_cluster[m].get(pid)
            rec[f"{m} [panel]"] = panel_pred[m].get(pid)
        subset_rows.append(rec)
    subset_df = pd.DataFrame(subset_rows)
    # Group recipes by their true label so per-cluster performance is scannable at a glance.
    subset_df = subset_df.sort_values(["M1_set", "M2_set", "Recipe"], ignore_index=True)

    # Per-true-cluster accuracy (rows = true cluster, columns = method) for M1 and M2.
    pc_m1 = per_cluster_accuracy(panel_pred, eval_truth, "M1_set")
    pc_m2 = per_cluster_accuracy(panel_pred, eval_truth, "M2_set")
    pc_m1_reach = per_cluster_accuracy(panel_pred, eval_truth, "M1_set", reachable_only=True)
    pc_m2_reach = per_cluster_accuracy(panel_pred, eval_truth, "M2_set", reachable_only=True)

    # Per-recipe predictions of ALL methods on the reachable recipes, methods best-first.
    reach = eval_truth[eval_truth["M1_set"].map(lambda s: bool(set(s) & REACHABLE))].copy()
    reach["_k1"] = reach["M1_set"].map(lambda s: ", ".join(s))
    reach["_k2"] = reach["M2_set"].map(lambda s: ", ".join(s))
    reach = reach.sort_values(["_k1", "_k2", "Recipe"])
    zrows = []
    for _, r in reach.iterrows():
        pid = r["pdm_id"]
        row = {"Recipe": r["Recipe"], "M1 true": ", ".join(r["M1_set"]), "M2 true": ", ".join(r["M2_set"])}
        for m in method_order:
            pred = panel_pred[m].get(pid)
            row[m] = None if (pred is None or isinstance(pred, float)) else str(pred)
        zrows.append(row)
    zmethods_reach = pd.DataFrame(zrows)

    def order_cols(df):
        return df[["True_cluster", "n"] + [m for m in method_order if m in df.columns]]

    pc_m1, pc_m2 = order_cols(pc_m1), order_cols(pc_m2)
    pc_m1_reach, pc_m2_reach = order_cols(pc_m1_reach), order_cols(pc_m2_reach)

    # All-recipe hand-over: raw MS argmax cluster per method (best-performing methods first).
    all_pred = pd.DataFrame({m: ms_cluster[m] for m in variants})[method_order]
    all_pred.index.name = "Rez.-Nr."

    truth_out = truth.drop(columns=["base"]).copy()
    truth_out["M1_set"] = truth_out["M1_set"].map(lambda s: ", ".join(s))
    truth_out["M2_set"] = truth_out["M2_set"].map(lambda s: ", ".join(s))

    # Readable column labels for the exported sheets (internal keys stay terse).
    READABLE = {
        "H": "Prozent %",
        "B_M1": "M1 Label Propagation",
        "C_M2": "M2 Rule Based",
        "D": "Cluster vorgegeben",
        "E": "Jan Free Sorting",
        "M1_set": "M1 true labels",
        "M2_set": "M2 true labels",
        "in_pdm": "In PDM export",
        "pdm_id": "PDM Rez.-Nr.",
        "evaluable": "Evaluable",
        "truth_method": "Truth method",
        "truth": "True label(s)",
    }
    truth_out = truth_out.rename(columns=READABLE)
    subset_df = subset_df.rename(columns=READABLE)
    excluded_df = excluded_df.rename(columns=READABLE)

    OUT_XLSX.parent.mkdir(exist_ok=True)
    with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as writer:
        pd.DataFrame(KEY_METHODS_EXPLANATION).to_excel(
            writer, sheet_name="key_Methods_Explanation", index=False)
        acc_m1.to_excel(writer, sheet_name="Accuracy_vs_M1", index=False)
        acc_m2.to_excel(writer, sheet_name="Accuracy_vs_M2", index=False)
        subset_df.to_excel(writer, sheet_name="Subset_Predictions", index=False)
        pc_m1.to_excel(writer, sheet_name="Per_Cluster_vs_M1", index=False)
        pc_m2.to_excel(writer, sheet_name="Per_Cluster_vs_M2", index=False)
        pc_m1_reach.to_excel(writer, sheet_name="Per_Cluster_vs_M1_reachable", index=False)
        pc_m2_reach.to_excel(writer, sheet_name="Per_Cluster_vs_M2_reachable", index=False)
        zmethods_reach.to_excel(writer, sheet_name="ZMethods_Reachable_17", index=False)
        truth_out.to_excel(writer, sheet_name="Truth_Labels", index=False)
        if len(excluded_df):
            excluded_df.to_excel(writer, sheet_name="Excluded_Recipes", index=False)
        all_pred.reset_index().to_excel(writer, sheet_name="All_Predictions_AllMethods", index=False)

    print(f"\nSaved: {OUT_XLSX}")
    print(f"\n=== Accuracy vs M1 (Label Propagation), {n_eval} recipes ===")
    print(acc_m1[["Method", "z_method", "Accuracy_%", "Accuracy_reachable_%"]].to_string(index=False))
    print(f"\n=== Accuracy vs M2 (Rule Based), {n_eval} recipes ===")
    print(acc_m2[["Method", "z_method", "Accuracy_%", "Accuracy_reachable_%"]].to_string(index=False))
    print(
        f"\nReachable-only (excludes dairy/Walderdbeere truths): "
        f"M1 keeps {n_keep_m1}/{n_eval}, M2 keeps {n_keep_m2}/{n_eval} recipes. "
        f"See Accuracy_reachable_% column and Excluded_Recipes sheet."
    )


if __name__ == "__main__":
    main()
